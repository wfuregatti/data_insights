import pandas as pd
import openpyxl

files = ['File1.xlsx', 'File2.xlsx']
result = pd.DataFrame()

for file in files:
    df = pd.read_excel(file)
    result = result.append(df)
result.to_excel('Result_file.xlsx')

workbook = openpyxl.load_workbook("Result_file.xlsx")
worksheet = workbook["Sheet1"]
worksheet.delete_cols(idx=1)
worksheet.delete_cols(idx=3)
worksheet.delete_cols(idx=6)
worksheet["D"+str(worksheet.max_row + 1)]="TOTAL SOLD"
worksheet["E"+str(worksheet.max_row)]="=SUM(E2:E41)"
workbook.save("Result_file.xlsx")
print("Result_file.xlsx Created")



